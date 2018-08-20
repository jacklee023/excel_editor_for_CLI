#!/bin/python3
import os
import sys
import logging
import argparse
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

                             # 中文字符         全角字符
pat_fullchar = re.compile(r'[\u4e00-\u9fa5]|[\u0391-\uffe5]') 
pat_chinese  = re.compile(r'[\u4e00-\u9fa5]')
pat_number   = re.compile(r'^\d+$')
pat_color    = re.compile("(.*│ )(=>\s*[^│])(\s*│.*)")
pat_color_sheet = re.compile("(.*)(\[.*\])(.*)")

not_need_lower_arg_ops_list = ['h', 'l', 'j', 'k', 'H', 'L', 'J', 'K']

glist_ops_quit        = ['quit'         , 'q']
glist_ops_setnum      = ['set_num'      , 'num']
glist_ops_fillcell    = ['fill'         , 'f']
glist_ops_hidden      = ['hidden'       , 'hid']
glist_ops_unhidden    = ['unhidden'     , 'unh']
glist_ops_createsheet = ['create'       , 'c']
glist_ops_removesheet = ['remove'       , 'r']
glist_ops_edidcell    = ['edit'         , 'e']
glist_ops_prevsheet   = ['prev'         , 'p']
glist_ops_nextsheet   = ['next'         , 'n']
glist_ops_moveleft    = ['moveleft'     , 'h']
glist_ops_moveright   = ['moveright'    , 'l']
glist_ops_movedown    = ['movedown'     , 'j']
glist_ops_moveup      = ['moveup'       , 'k']
glist_ops_shiftleft   = ['shiftleft'    , 'H']
glist_ops_shiftright  = ['shiftright'   , 'L']
glist_ops_shiftdown   = ['shiftdown'    , 'J']
glist_ops_shiftup     = ['shiftup'      , 'K']
glist_ops_setwidth    = ['setwidth'     , 'w']
glist_ops_show        = ['show'         , 's']
glist_ops_save        = ['save'         , ':w']
glist_ops_help        = ['help'         , 'dh']
glist_obj_row         = ['row'          , 'r']
glist_obj_col         = ['col'          , 'c']

command_info_list = [
    (glist_ops_quit         , 'quit this program'),
    (glist_ops_setnum       , 'set row/col number'),
    (glist_ops_fillcell     , 'fill whole row/col'),
    (glist_ops_hidden       , 'hidden curr row/col'),
    (glist_ops_unhidden     , 'unhidden curr row/col'),
    (glist_ops_createsheet  , 'create sheet'),
    (glist_ops_removesheet  , 'remove sheet'),
    (glist_ops_edidcell     , 'edit curr cell'),
    (glist_ops_prevsheet    , 'move to prev sheet'),
    (glist_ops_nextsheet    , 'move to next sheet'),
    (glist_ops_moveleft     , 'move the pointer to left'),
    (glist_ops_moveright    , 'move the pointer to right'),
    (glist_ops_movedown     , 'move the pointer to down'),
    (glist_ops_moveup       , 'move the pointer to up'),
    (glist_ops_shiftleft    , 'move the viewer to left'),
    (glist_ops_shiftright   , 'move the viewer to right'),
    (glist_ops_shiftdown    , 'move the viewer to down'),
    (glist_ops_shiftup      , 'move the viewer to up'),
    (glist_ops_setwidth     , 'set col width'),
    (glist_ops_show         , 'show curr view'),
    (glist_ops_save         , 'save curr table'),
    (glist_ops_help         , 'show this help message'),
]

def dump_log(args):
    abs_cur_path = os.path.abspath(os.path.expandvars(os.path.curdir))
    if args.debug:
        level = logging.DEBUG
    else :
        level = logging.INFO

    format_sh = '[%(levelname)s] %(message)s'
    #format_fh = '[%(levelname)s] %(message)s'
    format_fh = '%(asctime)s | %(funcName)s() | L%(lineno)s | [%(levelname)s] \n\t%(message)s'
    '''
        %(levelno)s     打印日志级别的数值
        %(levelname)s   打印日志级别名称
        %(pathname)s    打印当前执行程序的路径
        %(filename)s    打印当前执行程序名称
        %(funcName)s    打印日志的当前函数
        %(lineno)d      打印日志的当前行号
        %(asctime)s     打印日志的时间
        %(thread)d      打印线程id
        %(threadName)s  打印线程名称
        %(process)d     打印进程ID
        %(message)s     打印日志信息
    '''

    if args.log_file:

        logging.basicConfig(
            filename = args.log_file, 
            filemode = 'w', 
            datefmt  = "%y-%m-%d %H:%M:%S",
            format   = format_fh, 
            level    = level)

        console = logging.StreamHandler()
        console.setLevel(level)
        console.setFormatter(logging.Formatter(format_sh))
        logging.getLogger('').addHandler(console);
    else:
        logging.basicConfig(
            format = format_sh, 
            level  = level)

def get_args_top():
    parser = argparse.ArgumentParser()
    st = "debug option"
    parser.add_argument('-d',   '--debug', default = False, action='store_true', help=st)

    st = "set input excel file"
    parser.add_argument('-i',   '--excel_file', default = "./test.xlsx", help=st)

    st = "set output log file"
    parser.add_argument('-l',   '--log_file', default = "", help=st)

    args = parser.parse_args()
    return args

def clean_str(in_str, Null = False):
    if Null is True :
        return in_str
    elif in_str is None:
        return ""
    else :
        return str(in_str)

def get_wrap_strs(in_str, width):
    clean_str(in_str)

    in_str   += " " * (width + 1)
    wrap_str  = "" 
    wrap_strs = []
    str_len   = 0
    for i in range(len(in_str[:-1])) :
        curr_char = in_str[i]
        next_char = in_str[i+1]
        wrap_str += curr_char

        str_len += 2 if pat_fullchar.search(curr_char) != None else 1

        len_diff = width - str_len

        if len_diff == 1 and pat_fullchar.search(next_char) != None :
            wrap_str += " "
            flag = True
        elif len_diff == 0:
            flag = True
        else :
            flag = False

        if flag == True :
            wrap_strs.append(wrap_str)
            wrap_str = ""
            str_len = 0

    return wrap_strs

class excel_editer_for_cli(object):
    min_row = 2
    min_col = 2

    hchar  = "─"
    hbchar = "="
    vchar  = "│"
    mchar  = "┼"
    tchar  = "┬"
    bchar  = "┴"
    lchar  = "├"
    rchar  = "┤"
    ltchar = "┌"
    lbchar = "└"
    rtchar = "┐"
    rbchar = "┘"

    #hchar  = "-"
    #hbchar = "="
    #vchar  = "|"
    #mchar  = "+"
    #tchar  = "-"
    #bchar  = "-"
    #lchar  = "|"
    #rchar  = "|"
    #ltchar = "."
    #lbchar = "'"
    #rtchar = "."
    #rbchar = "'"

    def __init__(self, args, log, excel = ""):
        self.args = args
        self.log  = log
        self._hidden_rows = []
        self._hidden_cols = []

        self._col_num   = 4
        self._row_num   = 4
        self._col_width = 0
        self._col_widths = []
        self._str_row   = self.min_row
        self._str_col   = self.min_col
        self._end_row   = self._get_end_index(self._str_row, self._row_num, self._hidden_rows)
        self._end_col   = self._get_end_index(self._str_col, self._col_num, self._hidden_cols)
        self.excel      = excel

        self._curr_row  = self._str_row
        self._curr_col  = self._str_col
        self._curr_loc  = self._set_curr_loc()
        self._curr_index = 0

        self._open_excel()
        self.set_width(8)

    def _open_excel(self, read_only = False):
        self.workbook  = load_workbook(self.excel, read_only = read_only)
        self.worksheet = self.workbook.active

    def _sheet_move(self, step = 0):
        step += self._curr_index
        sheet_num = len(self.workbook.worksheets)
        while( step < 0) :
            step += sheet_num
        while( step >= sheet_num) :
            step -= sheet_num
        self.worksheet = self.workbook.worksheets[step]

    def _set_curr_loc(self, row=None, col=None):
        row = self._curr_row if row is None else row
        col = self._curr_col if col is None else col

        self._curr_loc = "%s%s"%(get_column_letter(col), row)
        return self._curr_loc

    def _get_unhidden_list(self, str_index, max_num, hidden_list, step = 1):
        unhidden_list = []
        index = str_index - 1
        cnt = 0
        while True :
            if cnt < max_num :
                print(cnt)
                index += step
                if cnt in hidden_list:
                    continue
                else:
                    cnt += 1
                unhidden_list.append(index)
            else :
                print('unhidden_list')
                return unhidden_list

    def _get_end_index(self, str_index, max_num, hidden_list, step = 1):
        # step = 1 / step = -1
        unhidden_list = self._get_unhidden_list(str_index, max_num, hidden_list, step)
        return unhidden_list[-1]

    def _set_view_range(self, str_row=None, str_col=None):
        st = []
        str_row = self._str_row if str_row is None else str_row
        str_col = self._str_col if str_col is None else str_col

        if str_row < self.min_row :
            str_row = self.min_row
            st.append("The start_row must be bigger than %s, and it has been reset to (%s)"%(self.min_row, self.min_row))
        if str_col < self.min_col :
            str_col = self.min_col
            st.append("The start_col must be bigger than %s, and it has been reset to (%s)"%(self.min_col, self.min_col))

        if str_row > self._curr_row :
            self._curr_row = str_row
        if str_col > self._curr_col :
            self._curr_col = str_col

        self._str_row = str_row 
        self._str_col = str_col
        self._end_row = self._get_end_index(self._str_row, self._row_num, self._hidden_rows)
        self._end_col = self._get_end_index(self._str_col, self._col_num, self._hidden_cols)

        return "\n".join(st)

    def _move_loc(self, row_step = 0, col_step = 0):
        st = []
        self._curr_row += row_step
        # skip hidden_rows
        while self._curr_row in self._hidden_rows:
            self._curr_row += 1 if row_step > 0 else -1
        if self._curr_row < self.min_row:
            self._curr_row = self._str_row
            st.append("The self._curr_row must be bigger than %s, and it has been reset to self._str_row(%s)"%(self.min_row, self._str_row))
        if self._curr_row < self._str_row :
            self.shift_up(-row_step)
        elif self._curr_row > self._end_row :
            self.shift_down(row_step)
        
        self._curr_col += col_step
        # skip hidden_cols
        while self._curr_col in self._hidden_cols:
            self._curr_col += 1 if col_step > 0 else -1
        if self._curr_col < self.min_col:
            self._curr_col = self._str_col
            st.append("The self._curr_col must be bigger than %s, and it has been reset to self._str_col(%s)"%(self.min_col, self._str_col))
        if self._curr_col < self._str_col :
            self.shift_left(-col_step)
        elif self._curr_col > self._end_col :
            self.shift_right(col_step)

        return "\n".join(st)

    def _get_table_data(self, ws):
        print('_get_table_data')
        self.show_data = []

        print(self._hidden_rows)
        print(self._hidden_cols)
        unhidden_rows = self._get_unhidden_list(self._str_row, self._row_num, self._hidden_rows)
        unhidden_cols = self._get_unhidden_list(self._str_col, self._col_num, self._hidden_cols)
        header_list   = [ "[%s]"%(self.worksheet.title) ]
        print(unhidden_rows)
        print(unhidden_cols)

        for i, row in enumerate(unhidden_rows) :
            cell_value = ws.cell(row = row, column = 1).value
            cell_value = clean_str(cell_value)
            col_name = "(%s)%s"%(row, cell_value)
            col_list = [ col_name ]
            for col in unhidden_cols :
                cell_value = ws.cell(row = 1, column = col).value
                cell_value = clean_str(cell_value)
                header_name = "(%s)%s"%(get_column_letter(col), cell_value)
                if header_name not in header_list :
                    header_list.append(header_name)
            
                cell_value = ws.cell(row = row, column = col).value
                cell_value = clean_str(cell_value)
                if row == self._curr_row and col == self._curr_col :
                    #cell_value =  "\033[1;32;40m %s \033[0m"%cell_value
                    col_list.append("=> %s"%cell_value)
                    #print("=> \033[1;32;40m %s \033[0m"%cell_value)
                    #self.log.info('\033[1;32;40m aaa \033[0m')
                else :
                    col_list.append(cell_value)
            self.show_data.append(col_list)

        self.show_data.insert(0, header_list)

        return self.show_data

    def _get_row_line(self, row_items, vchar = '│', widths = None):
        widths = self._col_widths if widths is None else widths

        line = "%s"%vchar
        for i, item in enumerate(row_items) :
            wrap_strs = get_wrap_strs(item, widths[i])
            line = "%s %s %s"%(line, wrap_strs[0], vchar)
        return line

    def _get_splitline(self, line, hchar = '─', vchar = '│', mchar = '┼', lchar = '├', rchar = '┤'):
        line = pat_chinese.sub(hchar * 2, line)
        line = re.sub(r"[%s]" %vchar, mchar, line)
        line = re.sub(r"[^%s]"%mchar, hchar, line)
        line = "%s%s%s"%(lchar,line[1:-1],rchar)
        return line 

    def _get_format_width(self):
        format_width = self._col_num
        for e in self._col_widths :
            format_width += e
            format_width += 2
        format_width -= 2

        return format_width

    def _get_cell_view(self, wrap=True):
        st_cell = []
        cell_value = self.worksheet.cell(row = self._curr_row, column = self._curr_col).value
        cell_value = clean_str(cell_value)

        line = "Cell Value : (%s) => %s"%(self._curr_loc, cell_value)

        format_width = self._get_format_width()
        cell_lines = get_wrap_strs(line, format_width) if wrap == True else [line]

        for cell_line in cell_lines :
            line = self._get_row_line(row_items = [cell_line], widths = [format_width], vchar = self.vchar)
            st_cell.append(line)

        splitline =       self._get_splitline(line,      hchar = self.hbchar, vchar = self.vchar, mchar = self.mchar, lchar = self.lchar,  rchar = self.rchar)
        st_cell.insert(0, self._get_splitline(splitline, hchar = self.hbchar, vchar = self.mchar, mchar = self.hchar, lchar = self.ltchar, rchar = self.rtchar) )
        st_cell.append(   self._get_splitline(splitline, hchar = self.hbchar, vchar = self.mchar, mchar = self.hchar, lchar = self.lbchar, rchar = self.rbchar) )

        return ("\n".join(st_cell))

    def _get_table_view(self):
        st_table = []
        for row in self.show_data :
            line = self._get_row_line(row, vchar = self.vchar)
            #print(line)
            #vchar  = "│"
            #m = pat_color.search(line)
            #if m is not None:
            #    prefix  = m.group(1)
            #    cell    = m.group(2)
            #    postfix = m.group(3)
            #    cell = "\033[1;32;40m%s\033[0m"%cell
            #    line = prefix + cell + postfix
            st_table.append(line)

        splitline        = self._get_splitline(line,      hchar = self.hchar, vchar = self.vchar, mchar = self.mchar, lchar = self.lchar,  rchar = self.rchar)
        st_table.insert(1, splitline)
        st_table.insert(0, self._get_splitline(splitline, hchar = self.hchar, vchar = self.mchar, mchar = self.tchar, lchar = self.ltchar, rchar = self.rtchar) )
        st_table.append(   self._get_splitline(splitline, hchar = self.hchar, vchar = self.mchar, mchar = self.bchar, lchar = self.lbchar, rchar = self.rbchar) )
        return("\n".join(st_table))

    def _get_sheet_view(self, wrap=True):
        st_sheet = []
        sheet_num  = len(self.workbook.sheetnames)
        line = ""
        for i, sheetname in enumerate(self.workbook.sheetnames):
            if self.workbook[sheetname] is self.worksheet :
                line = "%s [%s] "%(line, sheetname)
                #self.log.info("=> \033[1;32;40m %s \033[0m"%sheetname)
                self._curr_index = i
            else :
                line = "%s  %s "%(line, sheetname)

        line = "Sheet %s/%s %s"%(self._curr_index + 1, sheet_num, line)
                
        format_width = self._get_format_width()
        cell_lines = get_wrap_strs(line, format_width) if wrap == True else [line]
        for cell_line in cell_lines :
            line = self._get_row_line(row_items = [cell_line], widths = [format_width], vchar = self.vchar)
            #m = pat_color_sheet.search(line)
            #if m is not None:
            #    prefix  = m.group(1)
            #    cell    = m.group(2)
            #    postfix = m.group(3)
            #    cell = "\033[1;32;40m%s\033[0m"%cell
            #    print(cell)
            #    line = prefix + cell + postfix
            st_sheet.append(line)

        splitline =        self._get_splitline(line,      hchar = self.hbchar, vchar = self.vchar, mchar = self.mchar, lchar = self.lchar,  rchar = self.rchar)
        st_sheet.insert(0, self._get_splitline(splitline, hchar = self.hbchar, vchar = self.mchar, mchar = self.hchar, lchar = self.ltchar, rchar = self.rtchar) )
        st_sheet.append(   self._get_splitline(splitline, hchar = self.hbchar, vchar = self.mchar, mchar = self.hchar, lchar = self.lbchar, rchar = self.rbchar) )
        return("\n".join(st_sheet))

    def save(self, in_excel = None):
        excel = self.excel if in_excel is None else in_excel
        self.workbook.save(excel)

    def edit_cell(self, new_value):
        self.worksheet.cell(row = self._curr_row, column = self._curr_col, value = new_value)
    
    def sheet_prev(self, step = 1):
        self._sheet_move(-step)

    def sheet_next(self, step = 1):
        self._sheet_move(step)

    def create_sheet(self, sheetname, index = None):
        self.workbook.create_sheet(sheetname, index)
        self.worksheet = self.workbook[sheetname]

    def remove_sheet(self, sheetname):
        st = []
        if sheetname in self.workbook.sheetnames :
            rm_sheet = self.workbook[sheetname]
            self.workbook.remove(rm_sheet)
        else :
            st.append("sheet '%s' is not exist"%sheetname)

        return "\n".join(st)

    def fill(self, arg):
        if arg in glist_obj_row :
            self.fill_row()
        elif arg in glist_obj_col :
            self.fill_col()

    def fill_col(self, curr_col=None, min_row=None, max_row=None, fill_value=""):
        curr_col = self._curr_col          if curr_col is None else curr_col
        min_row  = self.worksheet.min_row if min_row  is None else min_row
        max_row  = self.worksheet.max_row if max_row  is None else max_row

        for row in range(min_row, max_row):
            self.worksheet.cell(row = row, column = curr_col, value = fill_value)

    def fill_row(self, curr_row=None, min_col=None, max_col=None, fill_value=""):
        curr_row = self._curr_row             if curr_row is None else curr_row
        min_col  = self.worksheet.min_column if min_col  is None else min_col
        max_col  = self.worksheet.max_column if max_col  is None else max_col

        for col in range(min_col, max_col):
            self.worksheet.cell(row = curr_row, column = col, value = fill_value)

    def set_width(self, width = None, index = None):
        if index is None :
            if width is not None :
                self._col_width = width

            self._col_widths = []
            for i in range(self._col_num + 1) :
                self._col_widths.append( self._col_width)
        else :
            self._col_widths[int(index)] = width

    def set_num(self, obj, arg):
        if obj in glist_obj_row :
            self.set_row_num(arg)
        elif obj in glist_obj_col :
            self.set_col_num(arg)

    def set_row_num(self, _row_num=None):
        self._row_num = _row_num if _row_num is not None else self._row_num
        self._end_row = self._get_end_index(self._str_row, self._row_num, self._hidden_rows)

    def set_col_num(self, _col_num=None):
        self._col_num = _col_num if _col_num is not None else self._col_num
        self._end_col = self._get_end_index(self._str_col, self._col_num, self._hidden_cols)

    def shift_left(self, step = 1):
        str_col = self._get_end_index(self._str_col + 1, step, self._hidden_cols, -1)
        self._set_view_range(str_col = str_col)

    def shift_right(self, step = 1):
        str_col = self._get_end_index(self._str_col + 1, step, self._hidden_cols, 1)
        self._set_view_range(str_col = str_col)

    def shift_up(self, step = 1):
        str_row = self._get_end_index(self._str_row + 1, step, self._hidden_rows, -1)
        self._set_view_range(str_row = str_row)

    def shift_down(self, step = 1):
        str_row = self._get_end_index(self._str_row + 1, step, self._hidden_rows, 1)
        self._set_view_range(str_row = str_row)

    def move_left(self, step = 1):
        self._move_loc(col_step = -step)

    def move_right(self, step = 1):
        self._move_loc(col_step = step)

    def move_up(self, step = 1):
        self._move_loc(row_step = -step)

    def move_down(self, step = 1):
        self._move_loc(row_step = step)

    def hidden(self, arg):
        if arg in glist_obj_row :
            print("before hidden_row")
            self.hidden_row()
            print("bbb")
        elif arg in glist_obj_col:
            self.hidden_col()
            print(arg)
        print("ccc")

    def hidden_row(self):
        if self._curr_row not in self._hidden_rows :
            self._hidden_rows.append(self._curr_row)
            self._curr_row += 1
            print("enter hidden_row")
        print("ddd")

        self.show()

    def hidden_col(self):
        if self._curr_col not in self._hidden_cols :
            self._hidden_cols.append(self._curr_col)
            self._curr_col += 1

    def unhidden(self, arg):
        if arg in glist_obj_row :
            self.unhidden_row()
        elif arg in glist_obj_col:
            self.unhidden_col()

    def unhidden_row(self, row):
        row = int(row)
        if row in self._hidden_rows :
            self._hidden_rows.remove(row)

    def unhidden_col(self, col):
        col = int(col)
        if col in self._hidden_cols :
            self._hidden_cols.remove(col)

    def show(self) :
        print('show 1')
        ws = self.worksheet

        self.set_width()
        print('show 2')
        self._set_curr_loc()
        print('show 2')
        self._get_table_data(ws)
        print('show 2')

        st_cell  = self._get_cell_view()
        st_table = self._get_table_view()
        st_sheet = self._get_sheet_view()
        print('show 3')

        st = []
        st.append("")
        st.append(st_cell)
        st.append(st_table)
        st.append(st_sheet)

        if len(self._hidden_rows) != 0 :
            st.append("hidden_rows = {} ".format( self._hidden_rows))
        if len(self._hidden_cols) != 0 :
            st.append("hidden_cols = {} ".format( self._hidden_cols))
        self.log.debug("str_row = {}, end_row = {}, curr_row = {} row_num = {}".format( self._str_row, self._end_row, self._curr_row, self._row_num))
        self.log.debug("str_col = {}, end_col = {}, curr_col = {} col_num = {}".format( self._str_col, self._end_col, self._curr_col, self._col_num))
        self.log.debug("str_char= {}, end_char= {}, curr_char= {} ".format( 
            get_column_letter(self._str_col), 
            get_column_letter(self._end_col), 
            get_column_letter(self._curr_col)))
        self.log.info("\n".join(st))

    def help(self):
        st = []
        st.append("Help Message")
        st.append("==============")
        for cmds, info in command_info_list:
            cmd = '/'.join(cmds)
            st.append("Command {:<20}, info : {}".format(cmd, info))
        #st.append("Use command q/quit to quit this program")
        st = '\n'.join(st)
        self.log.info(st)
        return st

    def interactive(self):
        pat_command = re.compile(r'^(\S+)\s+(\S+)\s+(\S+)|(\S+)\s+(\S+)|(\S+)$|^$')
        self.show()
        while True :
            (ops, obj, arg) = (None, None, None)
            self.log.info('Please enter command:')
            in_value = sys.stdin.readline().strip()
            match   = pat_command.search(in_value)
            opts    = match.groups()
            opt_cnt = len(opts) - opts.count(None)
            if opt_cnt == 3 :
                ops  = match.group(1)
                obj  = match.group(2)
                arg  = match.group(3)
            elif opt_cnt == 2 :
                ops  = match.group(4)
                arg  = match.group(5)
            elif opt_cnt == 1:
                ops  = match.group(6)
            elif opt_cnt == 0:
                self.log.warning("Empty command!")
                continue
            else :
                self.log.warning("opt_cnt is not 0,1,2 or 3")

            if arg is None :
                arg = 1
            elif pat_number.search(str(arg)) is not None :
                arg = int(arg)

            if ops not in not_need_lower_arg_ops_list :
                ops = ops.lower()

            if ops in glist_ops_edidcell and arg == 'None' :
                arg = ""

            if obj is not None :
                obj = obj.lower()

            self.log.debug("Command : ops={}, obj={}, arg={}".format(ops, obj, arg))
            if ops in glist_ops_quit:
                break
            elif ops in glist_ops_setnum:
                self.set_num(obj, arg)
            elif ops in glist_ops_fillcell:
                self.fill(arg)
            elif ops in glist_ops_hidden :
                self.hidden(arg)
                print('aaa')
            elif ops in glist_ops_unhidden :
                self.unhidden(arg)
            elif ops in glist_ops_createsheet :
                self.create_sheet(arg)
            elif ops in glist_ops_removesheet :
                self.remove_sheet(arg)
            elif ops in glist_ops_edidcell :
                self.edit_cell(arg)
            elif ops in glist_ops_prevsheet :
                self.sheet_prev(arg)
            elif ops in glist_ops_nextsheet :
                self.sheet_next(arg)
            elif ops in glist_ops_moveleft :
                self.move_left(arg)
            elif ops in glist_ops_moveright :
                self.move_right(arg)
            elif ops in glist_ops_movedown :
                self.move_down(arg)
            elif ops in glist_ops_moveup :
                self.move_up(arg)
            elif ops in glist_ops_shiftleft :
                self.shift_left(arg)
            elif ops in glist_ops_shiftright :
                self.shift_right(arg)
            elif ops in glist_ops_shiftdown :
                self.shift_down(arg)
            elif ops in glist_ops_shiftup :
                self.shift_up(arg)
            elif ops in glist_ops_setwidth :
                self.set_width(arg, obj)
            elif ops in glist_ops_show :
                self.show()
            elif ops in glist_ops_save :
                self.save(arg)
                break
            elif ops in glist_ops_help :
                self.help()
                continue
            else:
                self.log.warning("Command '{}' is not found! Please check help message :".format(in_value))
                self.help()
                continue

            self.show()

def main():
    args = get_args_top()
    dump_log(args)
    log = logging

    #excel_file = "./test.xlsx"
    excel_file = args.excel_file

    excel_obj = excel_editer_for_cli(args, log, excel_file)
    excel_obj.interactive()

if __name__ == '__main__':

    main()
